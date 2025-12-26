#!/usr/bin/env python3
"""
Taiwan 2025 Calendar Generator
Generates an Excel file with monthly grids, daily list, holiday markings, and automatic weekday calculations.
"""

import calendar
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Taiwan holidays for 2025
TAIWAN_HOLIDAYS_2025 = {
    (1, 1): "New Year's Day",
    (2, 8): "Lunar New Year Eve",
    (2, 9): "Lunar New Year",
    (2, 10): "Lunar New Year",
    (2, 11): "Lunar New Year",
    (2, 28): "Peace Memorial Day",
    (4, 4): "Children's Day",
    (4, 5): "Tomb Sweeping Day",
    (6, 10): "Dragon Boat Festival",
    (9, 17): "Mid-Autumn Festival",
    (10, 10): "Double Tenth Day",
    (12, 25): "Constitution Day",
}

# Weekday names
WEEKDAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
WEEKDAYS_SHORT = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

class TaiwanCalendar2025Generator:
    def __init__(self, filename="Taiwan_2025_Calendar.xlsx"):
        self.filename = filename
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet
        self.year = 2025
        
        # Define styles
        self.holiday_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        self.weekend_fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
        self.holiday_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
    def is_holiday(self, month, day):
        """Check if a date is a Taiwan holiday"""
        return (month, day) in TAIWAN_HOLIDAYS_2025
    
    def get_holiday_name(self, month, day):
        """Get the name of a holiday"""
        return TAIWAN_HOLIDAYS_2025.get((month, day), "")
    
    def is_weekend(self, weekday):
        """Check if a weekday is Saturday (5) or Sunday (6)"""
        return weekday in [5, 6]
    
    def create_monthly_grid_sheet(self, month):
        """Create a sheet with monthly grid view"""
        month_name = calendar.month_name[month]
        sheet_name = f"{month:02d}-{month_name}"
        ws = self.wb.create_sheet(sheet_name)
        
        # Set column widths
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # Title
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = f"Taiwan {self.year} - {month_name}"
        title_cell.font = Font(bold=True, size=16, color="1F4E78")
        title_cell.alignment = self.center_alignment
        
        # Weekday headers
        for col, day_name in enumerate(WEEKDAYS, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = day_name
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # Get calendar for the month
        month_calendar = calendar.monthcalendar(self.year, month)
        
        # Fill in the dates
        row = 4
        for week in month_calendar:
            for col, day in enumerate(week, 1):
                cell = ws.cell(row=row, column=col)
                if day != 0:
                    cell.value = day
                    cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
                    
                    # Apply holiday styling
                    if self.is_holiday(month, day):
                        cell.fill = self.holiday_fill
                        cell.font = self.holiday_font
                    # Apply weekend styling
                    elif self.is_weekend(col - 1):
                        cell.fill = self.weekend_fill
                    
                    cell.border = self.border
                    
                    # Set row height
                    ws.row_dimensions[row].height = 60
                else:
                    cell.border = self.border
            row += 1
        
        # Set row height for header
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[3].height = 20
    
    def create_daily_list_sheet(self):
        """Create a sheet with daily list"""
        ws = self.wb.create_sheet("Daily List", 0)
        
        # Set column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 20
        
        # Title
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = f"Taiwan {self.year} - Daily List"
        title_cell.font = Font(bold=True, size=14, color="1F4E78")
        title_cell.alignment = self.center_alignment
        ws.row_dimensions[1].height = 20
        
        # Headers
        headers = ["Date", "Weekday", "Holiday", "Notes"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        ws.row_dimensions[3].height = 20
        
        # Generate all dates for the year
        current_date = datetime(self.year, 1, 1)
        end_date = datetime(self.year, 12, 31)
        
        row = 4
        while current_date <= end_date:
            month = current_date.month
            day = current_date.day
            weekday = current_date.weekday()
            
            # Date column
            date_cell = ws.cell(row=row, column=1)
            date_cell.value = current_date.strftime("%Y-%m-%d")
            date_cell.alignment = Alignment(horizontal="center", vertical="center")
            date_cell.border = self.border
            
            # Weekday column
            weekday_cell = ws.cell(row=row, column=2)
            weekday_cell.value = WEEKDAYS[weekday]
            weekday_cell.alignment = self.center_alignment
            weekday_cell.border = self.border
            
            # Holiday column
            holiday_cell = ws.cell(row=row, column=3)
            if self.is_holiday(month, day):
                holiday_cell.value = self.get_holiday_name(month, day)
                holiday_cell.fill = self.holiday_fill
                holiday_cell.font = self.holiday_font
            elif self.is_weekend(weekday):
                holiday_cell.fill = self.weekend_fill
            
            holiday_cell.alignment = self.center_alignment
            holiday_cell.border = self.border
            
            # Notes column (empty for user input)
            notes_cell = ws.cell(row=row, column=4)
            notes_cell.border = self.border
            notes_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            
            current_date += timedelta(days=1)
            row += 1
        
        # Freeze panes
        ws.freeze_panes = "A4"
    
    def create_summary_sheet(self):
        """Create a summary sheet with holiday information"""
        ws = self.wb.create_sheet("Summary", 1)
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        
        # Title
        ws.merge_cells('A1:B1')
        title_cell = ws['A1']
        title_cell.value = f"Taiwan {self.year} - Holiday Summary"
        title_cell.font = Font(bold=True, size=14, color="1F4E78")
        title_cell.alignment = self.center_alignment
        ws.row_dimensions[1].height = 20
        
        # Headers
        date_header = ws['A3']
        date_header.value = "Date"
        date_header.fill = self.header_fill
        date_header.font = self.header_font
        date_header.alignment = self.center_alignment
        date_header.border = self.border
        
        holiday_header = ws['B3']
        holiday_header.value = "Holiday"
        holiday_header.fill = self.header_fill
        holiday_header.font = self.header_font
        holiday_header.alignment = self.center_alignment
        holiday_header.border = self.border
        
        ws.row_dimensions[3].height = 20
        
        # List all holidays
        row = 4
        for (month, day), holiday_name in sorted(TAIWAN_HOLIDAYS_2025.items()):
            date_cell = ws.cell(row=row, column=1)
            date_cell.value = f"{self.year}-{month:02d}-{day:02d}"
            date_cell.alignment = self.center_alignment
            date_cell.border = self.border
            date_cell.fill = self.holiday_fill
            date_cell.font = self.holiday_font
            
            holiday_cell = ws.cell(row=row, column=2)
            holiday_cell.value = holiday_name
            holiday_cell.alignment = Alignment(horizontal="left", vertical="center")
            holiday_cell.border = self.border
            holiday_cell.fill = self.holiday_fill
            holiday_cell.font = self.holiday_font
            
            row += 1
    
    def generate(self):
        """Generate the complete calendar workbook"""
        # Create daily list and summary first (they'll be reordered by sheet tabs)
        self.create_daily_list_sheet()
        self.create_summary_sheet()
        
        # Create monthly grid sheets
        for month in range(1, 13):
            self.create_monthly_grid_sheet(month)
        
        # Save the workbook
        self.wb.save(self.filename)
        print(f"✓ Calendar generated successfully: {self.filename}")


def main():
    """Main function to generate the Taiwan 2025 Calendar"""
    generator = TaiwanCalendar2025Generator()
    generator.generate()


if __name__ == "__main__":
    main()
